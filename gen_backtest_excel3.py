# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HDR_BG = "1A1A2E"; HDR_FG = "FFFFFF"
ALT_BG = "F9FAFB"; WIN_BG = "DCFCE7"; LOSE_BG = "FEF2F2"
POS_FG = "15803D"; NEG_FG = "DC2626"; BC = "DDDDDD"

def bd():
    s = Side(style='thin', color=BC)
    return Border(left=s, right=s, top=s, bottom=s)
def fl(h): return PatternFill("solid", fgColor=h)
def cn(): return Alignment(horizontal='center', vertical='center')
def lf(): return Alignment(horizontal='left', vertical='center')
def fn(bold=False, color="000000", sz=10):
    return Font(name='Arial', bold=bold, color=color, size=sz)

GEJU_COLOR = {
    "中和":   ("1D4ED8","DBEAFE"),
    "身弱":   ("92400E","FEF9C3"),
    "身强":   ("B91C1C","FEE2E2"),
    "从儿格": ("7C3AED","F3E8FF"),
    "从财格": ("7C3AED","F3E8FF"),
    "从官格": ("7C3AED","F3E8FF"),
}

# (name, gender, day_tg, geju, xi, ar_50, ar_0, bmr, alpha_50, alpha_0, mdd_50, mdd_0, sh_50, sh_0)
ALL_DATA = [
    ("啊",         "男","戊","从儿格","金、水",  0.023,0.024,0.022, 0.000, 0.002,-0.256,-0.256,0.25,0.27),
    ("Cq女",       "女","辛","身弱",  "土、金",  0.032,0.028,0.022, 0.010, 0.006,-0.256,-0.256,0.32,0.29),
    ("程钰",       "女","丙","身弱",  "木、火",  0.023,0.017,0.022, 0.001,-0.006,-0.216,-0.197,0.30,0.26),
    ("程远",       "男","甲","身弱",  "水、木",  0.002,-0.001,0.022,-0.020,-0.023,-0.117,-0.079,0.08,-0.07),
    ("程宇宸",     "女","辛","中和",  "木、火、土",0.040,0.040,0.022,0.018,0.018,-0.289,-0.289,0.37,0.38),
    ("传奇妹夫",   "男","己","身弱",  "火、土",  0.019,0.006,0.022,-0.003,-0.016,-0.115,-0.079,0.38,0.19),
    ("CPR",        "女","甲","身弱",  "水、木",  0.006,0.000,0.022,-0.016,-0.022,-0.154,-0.125,0.15,0.01),
    ("Cq男",       "男","壬","身强",  "火、土",  0.028,0.020,0.022, 0.005,-0.002,-0.226,-0.226,0.32,0.26),
    ("代",         "男","壬","中和",  "火、土、金",0.041,0.042,0.022,0.019,0.019,-0.287,-0.285,0.37,0.37),
    ("戴常",       "男","癸","身弱",  "金、水",  0.011,0.009,0.022,-0.012,-0.013,-0.218,-0.200,0.18,0.17),
    ("戴元媛",     "女","己","身弱",  "火、土",  0.028,0.021,0.022, 0.006,-0.001,-0.170,-0.137,0.37,0.35),
    ("大珂国学",   "女","丙","中和",  "金、水、木",0.042,0.042,0.022,0.020,0.020,-0.289,-0.289,0.37,0.38),
    ("大侄子对象", "女","庚","中和",  "木、火、土",0.039,0.038,0.022,0.017,0.016,-0.286,-0.284,0.37,0.37),
    ("大侄子毅",   "男","庚","中和",  "木、火、土",0.041,0.041,0.022,0.019,0.019,-0.289,-0.289,0.37,0.37),
    ("东东",       "男","戊","中和",  "水、木、火",0.042,0.042,0.022,0.020,0.020,-0.277,-0.272,0.38,0.38),
    ("Fl",         "男","戊","中和",  "水、木、火",0.042,0.043,0.022,0.020,0.021,-0.269,-0.257,0.38,0.39),
    ("郭艾伦",     "男","己","从财格","水、木",  0.005,-0.002,0.022,-0.017,-0.024,-0.195,-0.158,0.11,-0.01),
    ("黄日华",     "女","甲","从儿格","火、土",  0.029,0.022,0.022, 0.007, 0.000,-0.255,-0.231,0.32,0.27),
    ("胡姐",       "女","癸","身弱",  "金、水",  0.020,0.013,0.022,-0.002,-0.010,-0.247,-0.247,0.27,0.20),
    ("九龙道长",   "男","丁","中和",  "金、水、木",0.042,0.042,0.022,0.020,0.020,-0.289,-0.289,0.37,0.38),
    ("Kylie",      "女","癸","从官格","土、火",  0.031,0.024,0.022, 0.009, 0.001,-0.226,-0.187,0.39,0.37),
    ("李苏涛",     "男","壬","身弱",  "金、水",  0.013,0.008,0.022,-0.009,-0.014,-0.247,-0.247,0.19,0.14),
    ("柳忠祥",     "男","癸","身强",  "火、土",  0.015,0.010,0.022,-0.007,-0.012,-0.158,-0.151,0.29,0.23),
    ("Lxt",        "女","壬","中和",  "火、土、金",0.041,0.042,0.022,0.019,0.019,-0.287,-0.285,0.37,0.37),
    ("马为",       "女","己","身弱",  "火、土",  0.028,0.017,0.022, 0.006,-0.005,-0.170,-0.137,0.38,0.30),
    ("么么撒",     "女","乙","身强",  "土、金",  0.029,0.023,0.022, 0.007, 0.001,-0.210,-0.177,0.35,0.33),
    ("宋丽娜",     "女","丙","中和",  "金、水、木",0.041,0.041,0.022,0.019,0.019,-0.289,-0.289,0.37,0.37),
    ("孙洁",       "女","丙","身弱",  "木、火",  0.028,0.024,0.022, 0.006, 0.002,-0.216,-0.197,0.35,0.35),
    ("孙平恒大",   "女","癸","中和",  "火、土、金",0.034,0.030,0.022,0.012,0.008,-0.254,-0.231,0.34,0.33),
    ("孙倩倩",     "女","辛","身弱",  "土、金",  0.023,0.018,0.022, 0.001,-0.004,-0.197,-0.160,0.30,0.28),
    ("谭鑫龙",     "男","丙","身强",  "金、水",  0.028,0.028,0.022, 0.006, 0.006,-0.256,-0.256,0.31,0.32),
    ("特朗普",     "男","己","身强",  "水、木",  0.009,0.008,0.022,-0.013,-0.015,-0.162,-0.133,0.18,0.18),
    ("Tsq",        "男","壬","中和",  "火、土、金",0.041,0.042,0.022,0.019,0.019,-0.287,-0.285,0.37,0.37),
    ("小沈阳",     "男","乙","从儿格","火、土",  0.038,0.037,0.022, 0.016, 0.015,-0.226,-0.226,0.40,0.41),
    ("圆圆爸爸",   "男","丁","中和",  "金、水、木",0.042,0.042,0.022,0.020,0.020,-0.289,-0.289,0.37,0.38),
    ("圆圆妈妈",   "女","丙","身弱",  "木、火",  0.016,0.011,0.022,-0.006,-0.011,-0.199,-0.160,0.28,0.24),
    ("圆圆姐",     "女","癸","从官格","土、火",  0.033,0.030,0.022, 0.011, 0.008,-0.255,-0.231,0.36,0.36),
    ("圆圆姨夫",   "男","壬","中和",  "火、土、金",0.034,0.030,0.022,0.012,0.008,-0.254,-0.231,0.34,0.33),
    ("张富跃",     "男","庚","身弱",  "土、金",  0.031,0.022,0.022, 0.009, 0.000,-0.197,-0.160,0.38,0.33),
    ("张丽",       "女","庚","身弱",  "土、金",  0.018,0.013,0.022,-0.004,-0.009,-0.179,-0.152,0.26,0.25),
    ("张松",       "女","庚","中和",  "木、火、土",0.041,0.041,0.022,0.019,0.019,-0.289,-0.289,0.37,0.37),
    ("张文",       "男","壬","身弱",  "金、水",  0.006,0.005,0.022,-0.016,-0.017,-0.189,-0.153,0.14,0.14),
    ("张跃芳",     "女","丁","中和",  "金、水、木",0.041,0.041,0.022,0.019,0.019,-0.289,-0.289,0.37,0.37),
    ("郑会杰",     "男","戊","中和",  "水、木、火",0.042,0.043,0.022,0.020,0.021,-0.269,-0.257,0.38,0.39),
    ("钟声",       "女","庚","中和",  "木、火、土",0.040,0.040,0.022,0.018,0.018,-0.289,-0.289,0.37,0.38),
    ("周勃舒",     "男","辛","身弱",  "土、金",  0.027,0.022,0.022, 0.005, 0.000,-0.139,-0.128,0.37,0.39),
    ("周顾",       "男","丙","从儿格","土、金",  0.037,0.034,0.022, 0.015, 0.012,-0.222,-0.220,0.37,0.36),
]

def make_excel(filepath, title_line, note_line, ar_idx, alpha_idx, mdd_idx, sh_idx):
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    # 标题（不合并，避免排序冲突）
    c = ws["A1"]
    c.value = title_line
    c.font = Font(name='Arial', bold=True, color="1A1A2E", size=12)
    c.fill = fl("EFF6FF"); c.alignment = lf()
    ws.row_dimensions[1].height = 28

    c = ws["A2"]
    c.value = note_line
    c.font = Font(name='Arial', color="555555", size=9)
    c.fill = fl("F8FAFF"); c.alignment = lf()
    ws.row_dimensions[2].height = 16

    ws.row_dimensions[3].height = 6

    # 表头
    headers = ["#","命主","性别","日主","格局","喜用","策略年化","基准年化","超额","最大回撤","夏普","跑赢"]
    ws.row_dimensions[4].height = 22
    for ci, hdr in enumerate(headers, 1):
        c = ws.cell(row=4, column=ci, value=hdr)
        c.font = Font(name='Arial', bold=True, color=HDR_FG, size=10)
        c.fill = fl(HDR_BG); c.alignment = cn(); c.border = bd()

    # 自动筛选（排序）
    ws.auto_filter.ref = "A4:L{}".format(4 + len(ALL_DATA))

    # 数据行
    for i, row in enumerate(ALL_DATA):
        r = i + 5
        name, gender, day_tg, geju, xi = row[0], row[1], row[2], row[3], row[4]
        ar     = row[ar_idx]
        bmr    = row[7]   # index 7 = bmr，始终为0.022
        alpha  = row[alpha_idx]
        mdd    = row[mdd_idx]
        sharpe = row[sh_idx]
        win    = alpha > 0.0005
        lose   = alpha < -0.0005
        base   = "FFFFFF" if i % 2 == 0 else ALT_BG
        row_bg = WIN_BG if win else (LOSE_BG if lose else base)

        ws.row_dimensions[r].height = 18

        def wc(ci, val, nf=None, bold=False, fg="000000", bg=None):
            c = ws.cell(row=r, column=ci, value=val)
            c.font = Font(name='Arial', bold=bold, color=fg, size=10)
            c.fill = fl(bg or base)
            c.alignment = cn()
            c.border = bd()
            if nf: c.number_format = nf
            return c

        wc(1, i+1, fg="888888")
        c2 = wc(2, name, bold=True, fg="1A1A2E")
        c2.alignment = lf()
        wc(3, gender)
        wc(4, day_tg, bold=True)

        # 格局
        gc, gb = GEJU_COLOR.get(geju, ("333333", base))
        c = ws.cell(row=r, column=5, value=geju)
        c.font = Font(name='Arial', bold=True, color=gc, size=9)
        c.fill = fl(gb); c.alignment = cn(); c.border = bd()

        c6 = wc(6, xi)
        c6.alignment = lf()

        ar_fg = POS_FG if ar >= 0.022 else (NEG_FG if ar < 0.015 else "000000")
        wc(7, ar, nf="0.0%", bold=True, fg=ar_fg)
        wc(8, bmr, nf="0.0%", fg="555555")

        al_fg = POS_FG if win else (NEG_FG if lose else "888888")
        wc(9, alpha, nf="+0.0%;-0.0%;0.0%", bold=True, fg=al_fg, bg=row_bg)
        wc(10, mdd, nf="0.0%", fg=NEG_FG)

        sh_fg = POS_FG if sharpe >= 0.35 else ("000000" if sharpe >= 0.2 else NEG_FG)
        wc(11, sharpe, nf="0.00", fg=sh_fg)

        win_val = 1 if win else (0 if lose else 0)
        c = ws.cell(row=r, column=12, value=win_val)
        c.number_format = '"跑赢";;"-"'
        c.font = Font(name='Arial', bold=True, color=POS_FG if win else NEG_FG, size=10)
        c.fill = fl(WIN_BG if win else (LOSE_BG if lose else base))
        c.alignment = cn(); c.border = bd()

    # 汇总行（不合并）
    last = 5 + len(ALL_DATA)
    ws.row_dimensions[last].height = 22
    c = ws.cell(row=last, column=1, value="平均 / 汇总")
    c.font = Font(name='Arial', bold=True, color=HDR_FG, size=10)
    c.fill = fl(HDR_BG); c.alignment = lf(); c.border = bd()

    for ci, nf, col_letter in [
        (7,  "0.0%",             "G"),
        (8,  "0.0%",             "H"),
        (9,  "+0.0%;-0.0%;0.0%", "I"),
        (10, "0.0%",             "J"),
        (11, "0.00",             "K"),
    ]:
        c = ws.cell(row=last, column=ci,
                    value="=AVERAGE({}5:{}{})" .format(col_letter, col_letter, last-1))
        c.number_format = nf
        c.font = Font(name='Arial', bold=True, color=HDR_FG, size=10)
        c.fill = fl(HDR_BG); c.alignment = cn(); c.border = bd()

    # 胜率 = SUM of win column
    c = ws.cell(row=last, column=12,
                value="=SUM(L5:L{})".format(last-1))
    c.number_format = '0"/47 胜率"'
    c.font = Font(name='Arial', bold=True, color=HDR_FG, size=10)
    c.fill = fl(HDR_BG); c.alignment = cn(); c.border = bd()

    # 列宽
    widths = [4, 13, 5, 5, 9, 15, 9, 9, 9, 9, 7, 8]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A5"
    wb.save(filepath)
    print("Saved:", filepath)

make_excel(
    r"c:\Users\tgspc\bazi-app\backtest_纯评分仓位.xlsx",
    "八字量化回测 2015-2025  |  沪深300ETF  |  纯评分仓位（无底仓）",
    "仓位：评分>=70=100%  55~69=60%  45~54=30%  <45=空仓  |  大盘过滤：沪深300>120日均线  |  月调仓 万3手续费",
    ar_idx=6, alpha_idx=9, mdd_idx=11, sh_idx=13
)

make_excel(
    r"c:\Users\tgspc\bazi-app\backtest_底仓50.xlsx",
    "八字量化回测 2015-2025  |  沪深300ETF  |  底仓50%（评分加减仓）",
    "仓位：评分>=70=100%  55~69=75%  45~54=50%(底仓)  35~44=25%  <35=空仓  |  大盘过滤：沪深300>120日均线",
    ar_idx=5, alpha_idx=8, mdd_idx=10, sh_idx=12
)
