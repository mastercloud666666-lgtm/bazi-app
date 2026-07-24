from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00

wb = Workbook()
ws = wb.active
ws.title = "回测结果"

# ── 配色 ────────────────────────────────────────────────
HDR_BG   = "1A1A2E"
HDR_FG   = "FFFFFF"
META_BG  = "F0F4FF"
POS_FG   = "15803D"  # green
NEG_FG   = "DC2626"  # red
ALT_BG   = "F9FAFB"
WIN_BG   = "DCFCE7"
LOSE_BG  = "FEF2F2"
BORDER_C = "DDDDDD"

def side(c=BORDER_C): return Side(style='thin', color=c)
def border(): return Border(left=side(), right=side(), top=side(), bottom=side())
def hdr_font(sz=10): return Font(name='微软雅黑', bold=True, color=HDR_FG, size=sz)
def cell_font(bold=False, color="000000", sz=10):
    return Font(name='微软雅黑', bold=bold, color=color, size=sz)
def fill(hex_color): return PatternFill("solid", fgColor=hex_color)
def center(): return Alignment(horizontal='center', vertical='center', wrap_text=False)
def left():   return Alignment(horizontal='left',   vertical='center')

# ── 标题区 ───────────────────────────────────────────────
ws.merge_cells("A1:L1")
ws["A1"] = "八字量化回测结果  |  2015-01-01 ~ 2025-12-31  |  沪深300ETF（510300）仓位择时"
ws["A1"].font = Font(name='微软雅黑', bold=True, color="1A1A2E", size=13)
ws["A1"].fill = fill("EFF6FF")
ws["A1"].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 28

ws.merge_cells("A2:L2")
ws["A2"] = "仓位阶梯：评分≥70=100%   55~69=60%   45~54=30%   <45=空仓    大盘过滤：沪深300>120日均线    月调仓，万3手续费"
ws["A2"].font = Font(name='微软雅黑', color="555555", size=9)
ws["A2"].fill = fill("F8FAFF")
ws["A2"].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[2].height = 18

# ── 统计摘要行 ────────────────────────────────────────────
ws.merge_cells("A3:B3"); ws["A3"] = "命主总数";  ws["A3"].alignment=center()
ws["C3"] = 47
ws.merge_cells("D3:E3"); ws["D3"] = "跑赢基准";  ws["D3"].alignment=center()
ws["F3"] = "=COUNTIF(K6:K52,\"✅\")"
ws.merge_cells("G3:H3"); ws["G3"] = "胜率";       ws["G3"].alignment=center()
ws["I3"] = "=F3/C3"
ws["I3"].number_format = "0%"
ws.merge_cells("J3:K3"); ws["J3"] = "基准年化";    ws["J3"].alignment=center()
ws["L3"] = 0.022
ws["L3"].number_format = "0.0%"

for col in ["A3","D3","G3","J3"]:
    ws[col].font = Font(name='微软雅黑', bold=True, color="1A1A2E", size=10)
    ws[col].fill = fill("DBEAFE")
for col in ["C3","F3","I3","L3"]:
    ws[col].font = Font(name='微软雅黑', bold=True, color="1A1A2E", size=11)
    ws[col].fill = fill("EFF6FF")
ws.row_dimensions[3].height = 22

ws.row_dimensions[4].height = 6  # spacer

# ── 表头 ─────────────────────────────────────────────────
headers = ["#","命主","性别","日主","格局","喜用","策略年化","基准年化","超额","最大回撤","夏普","跑赢"]
cols    = ["A","B","C","D","E","F","G","H","I","J","K","L"]
ws.row_dimensions[5].height = 22

for col, hdr in zip(cols, headers):
    c = ws[f"{col}5"]
    c.value = hdr
    c.font = hdr_font()
    c.fill = fill(HDR_BG)
    c.alignment = center()
    c.border = border()

# ── 数据 ─────────────────────────────────────────────────
rows = [
    ("啊",         "男","戊","从儿格", "金、水",       0.024, 0.022,  0.002, -0.256, 0.27),
    ("Cq女",       "女","辛","身弱",   "土、金",       0.028, 0.022,  0.006, -0.256, 0.29),
    ("程钰",       "女","丙","身弱",   "木、火",       0.017, 0.022, -0.006, -0.197, 0.26),
    ("程远",       "男","甲","身弱",   "水、木",      -0.001, 0.022, -0.023, -0.079,-0.07),
    ("程宇宸",     "女","辛","中和",   "木、火、土",   0.040, 0.022,  0.018, -0.289, 0.38),
    ("传奇妹夫",   "男","己","身弱",   "火、土",       0.006, 0.022, -0.016, -0.079, 0.19),
    ("CPR",        "女","甲","身弱",   "水、木",       0.000, 0.022, -0.022, -0.125, 0.01),
    ("Cq男",       "男","壬","身强",   "火、土",       0.020, 0.022, -0.002, -0.226, 0.26),
    ("代",         "男","壬","中和",   "火、土、金",   0.042, 0.022,  0.019, -0.285, 0.37),
    ("戴常",       "男","癸","身弱",   "金、水",       0.009, 0.022, -0.013, -0.200, 0.17),
    ("戴元媛",     "女","己","身弱",   "火、土",       0.021, 0.022, -0.001, -0.137, 0.35),
    ("大珂国学",   "女","丙","中和",   "金、水、木",   0.042, 0.022,  0.020, -0.289, 0.38),
    ("大侄子对象", "女","庚","中和",   "木、火、土",   0.038, 0.022,  0.016, -0.284, 0.37),
    ("大侄子毅",   "男","庚","中和",   "木、火、土",   0.041, 0.022,  0.019, -0.289, 0.37),
    ("东东",       "男","戊","中和",   "水、木、火",   0.042, 0.022,  0.020, -0.272, 0.38),
    ("Fl",         "男","戊","中和",   "水、木、火",   0.043, 0.022,  0.021, -0.257, 0.39),
    ("郭艾伦",     "男","己","从财格", "水、木",      -0.002, 0.022, -0.024, -0.158,-0.01),
    ("黄日华",     "女","甲","从儿格", "火、土",       0.022, 0.022,  0.000, -0.231, 0.27),
    ("胡姐",       "女","癸","身弱",   "金、水",       0.013, 0.022, -0.010, -0.247, 0.20),
    ("九龙道长",   "男","丁","中和",   "金、水、木",   0.042, 0.022,  0.020, -0.289, 0.38),
    ("Kylie",      "女","癸","从官格", "土、火",       0.024, 0.022,  0.001, -0.187, 0.37),
    ("李苏涛",     "男","壬","身弱",   "金、水",       0.008, 0.022, -0.014, -0.247, 0.14),
    ("柳忠祥",     "男","癸","身强",   "火、土",       0.010, 0.022, -0.012, -0.151, 0.23),
    ("Lxt",        "女","壬","中和",   "火、土、金",   0.042, 0.022,  0.019, -0.285, 0.37),
    ("马为",       "女","己","身弱",   "火、土",       0.017, 0.022, -0.005, -0.137, 0.30),
    ("么么撒",     "女","乙","身强",   "土、金",       0.023, 0.022,  0.001, -0.177, 0.33),
    ("宋丽娜",     "女","丙","中和",   "金、水、木",   0.041, 0.022,  0.019, -0.289, 0.37),
    ("孙洁",       "女","丙","身弱",   "木、火",       0.024, 0.022,  0.002, -0.197, 0.35),
    ("孙平恒大",   "女","癸","中和",   "火、土、金",   0.030, 0.022,  0.008, -0.231, 0.33),
    ("孙倩倩",     "女","辛","身弱",   "土、金",       0.018, 0.022, -0.004, -0.160, 0.28),
    ("谭鑫龙",     "男","丙","身强",   "金、水",       0.028, 0.022,  0.006, -0.256, 0.32),
    ("特朗普",     "男","己","身强",   "水、木",       0.008, 0.022, -0.015, -0.133, 0.18),
    ("Tsq",        "男","壬","中和",   "火、土、金",   0.042, 0.022,  0.019, -0.285, 0.37),
    ("小沈阳",     "男","乙","从儿格", "火、土",       0.037, 0.022,  0.015, -0.226, 0.41),
    ("圆圆爸爸",   "男","丁","中和",   "金、水、木",   0.042, 0.022,  0.020, -0.289, 0.38),
    ("圆圆妈妈",   "女","丙","身弱",   "木、火",       0.011, 0.022, -0.011, -0.160, 0.24),
    ("圆圆姐",     "女","癸","从官格", "土、火",       0.030, 0.022,  0.008, -0.231, 0.36),
    ("圆圆姨夫",   "男","壬","中和",   "火、土、金",   0.030, 0.022,  0.008, -0.231, 0.33),
    ("张富跃",     "男","庚","身弱",   "土、金",       0.022, 0.022,  0.000, -0.160, 0.33),
    ("张丽",       "女","庚","身弱",   "土、金",       0.013, 0.022, -0.009, -0.152, 0.25),
    ("张松",       "女","庚","中和",   "木、火、土",   0.041, 0.022,  0.019, -0.289, 0.37),
    ("张文",       "男","壬","身弱",   "金、水",       0.005, 0.022, -0.017, -0.153, 0.14),
    ("张跃芳",     "女","丁","中和",   "金、水、木",   0.041, 0.022,  0.019, -0.289, 0.37),
    ("郑会杰",     "男","戊","中和",   "水、木、火",   0.043, 0.022,  0.021, -0.257, 0.39),
    ("钟声",       "女","庚","中和",   "木、火、土",   0.040, 0.022,  0.018, -0.289, 0.38),
    ("周勃舒",     "男","辛","身弱",   "土、金",       0.022, 0.022,  0.000, -0.128, 0.39),
    ("周顾",       "男","丙","从儿格", "土、金",       0.034, 0.022,  0.012, -0.220, 0.36),
]

for i, row in enumerate(rows):
    r = i + 6  # Excel row
    even = (i % 2 == 0)
    base_bg = "FFFFFF" if not even else ALT_BG
    name, gender, day_tg, geju, xi, ar, bmr, alpha, mdd, sharpe = row
    win = alpha > 0
    row_bg = WIN_BG if win else (LOSE_BG if alpha < 0 else base_bg)

    ws.row_dimensions[r].height = 18

    def wc(col, val, num_fmt=None, bold=False, color=None, bg=None, align=None):
        c = ws[f"{col}{r}"]
        c.value = val
        c.font = cell_font(bold=bold, color=color or "000000")
        c.fill = fill(bg or base_bg)
        c.alignment = align or center()
        c.border = border()
        if num_fmt: c.number_format = num_fmt

    wc("A", i+1, bold=False, color="888888")
    wc("B", name, bold=True,  color="1A1A2E", align=left())
    wc("C", gender)
    wc("D", day_tg, bold=True)

    # 格局 with color
    geju_colors = {"中和":"1D4ED8","身弱":"92400E","身强":"B91C1C","从儿格":"7C3AED","从财格":"7C3AED","从官格":"7C3AED"}
    geju_bgs    = {"中和":"DBEAFE","身弱":"FEF9C3","身强":"FEE2E2","从儿格":"F3E8FF","从财格":"F3E8FF","从官格":"F3E8FF"}
    gj_color = geju_colors.get(geju, "333333")
    gj_bg    = geju_bgs.get(geju, base_bg)
    c = ws[f"E{r}"]
    c.value = geju
    c.font = Font(name='微软雅黑', bold=True, color=gj_color, size=9)
    c.fill = fill(gj_bg)
    c.alignment = center()
    c.border = border()

    wc("F", xi, align=left())

    # 策略年化
    ar_color = POS_FG if ar >= 0.022 else (NEG_FG if ar < 0.015 else "000000")
    wc("G", ar, num_fmt="0.0%", bold=True, color=ar_color)
    wc("H", bmr, num_fmt="0.0%", color="555555")

    # 超额
    al_color = POS_FG if alpha > 0 else (NEG_FG if alpha < 0 else "888888")
    wc("I", alpha, num_fmt="+0.0%;-0.0%;0.0%", bold=True, color=al_color, bg=row_bg)

    # 最大回撤
    wc("J", mdd, num_fmt="0.0%", color=NEG_FG)

    # 夏普
    sh_color = POS_FG if sharpe >= 0.35 else ("000000" if sharpe >= 0.2 else NEG_FG)
    wc("K", sharpe, num_fmt="0.00", color=sh_color)

    # 跑赢
    c = ws[f"L{r}"]
    c.value = "✅" if win else ("—" if alpha == 0 else "❌")
    c.font = Font(name='微软雅黑', size=11)
    c.fill = fill(WIN_BG if win else (LOSE_BG if alpha < 0 else base_bg))
    c.alignment = center()
    c.border = border()

# ── 汇总行 ───────────────────────────────────────────────
last = len(rows) + 6  # row after data
ws.row_dimensions[last].height = 20
ws.merge_cells(f"A{last}:F{last}")
ws[f"A{last}"] = f"汇总  |  47人  |  2015-2025  |  沪深300ETF择时  |  纯评分仓位（无底仓）"
ws[f"A{last}"].font = Font(name='微软雅黑', bold=True, color=HDR_FG, size=9)
ws[f"A{last}"].fill = fill(HDR_BG)
ws[f"A{last}"].alignment = left()
ws[f"A{last}"].border = border()

ws[f"G{last}"] = f'=AVERAGE(G6:G{last-1})'
ws[f"G{last}"].number_format = "0.0%"
ws[f"G{last}"].font = Font(name='微软雅黑', bold=True, color=HDR_FG)
ws[f"G{last}"].fill = fill(HDR_BG)
ws[f"G{last}"].alignment = center()
ws[f"G{last}"].border = border()

ws[f"H{last}"] = 0.022
ws[f"H{last}"].number_format = "0.0%"
ws[f"H{last}"].font = Font(name='微软雅黑', bold=True, color=HDR_FG)
ws[f"H{last}"].fill = fill(HDR_BG)
ws[f"H{last}"].alignment = center()
ws[f"H{last}"].border = border()

ws[f"I{last}"] = f'=AVERAGE(I6:I{last-1})'
ws[f"I{last}"].number_format = "+0.0%;-0.0%;0.0%"
ws[f"I{last}"].font = Font(name='微软雅黑', bold=True, color=HDR_FG)
ws[f"I{last}"].fill = fill(HDR_BG)
ws[f"I{last}"].alignment = center()
ws[f"I{last}"].border = border()

ws[f"J{last}"] = f'=AVERAGE(J6:J{last-1})'
ws[f"J{last}"].number_format = "0.0%"
ws[f"J{last}"].font = Font(name='微软雅黑', bold=True, color=HDR_FG)
ws[f"J{last}"].fill = fill(HDR_BG)
ws[f"J{last}"].alignment = center()
ws[f"J{last}"].border = border()

ws[f"K{last}"] = f'=AVERAGE(K6:K{last-1})'
ws[f"K{last}"].number_format = "0.00"
ws[f"K{last}"].font = Font(name='微软雅黑', bold=True, color=HDR_FG)
ws[f"K{last}"].fill = fill(HDR_BG)
ws[f"K{last}"].alignment = center()
ws[f"K{last}"].border = border()

ws[f"L{last}"] = f'=COUNTIF(L6:L{last-1},"✅")&"/47  ("&TEXT(COUNTIF(L6:L{last-1},"✅")/47,"0%")&")"'
ws[f"L{last}"].font = Font(name='微软雅黑', bold=True, color=HDR_FG)
ws[f"L{last}"].fill = fill(HDR_BG)
ws[f"L{last}"].alignment = center()
ws[f"L{last}"].border = border()

# ── 列宽 ─────────────────────────────────────────────────
widths = {"A":5,"B":13,"C":6,"D":6,"E":9,"F":14,"G":9,"H":9,"I":9,"J":9,"K":7,"L":10}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# ── 冻结首行区域 ──────────────────────────────────────────
ws.freeze_panes = "A6"

out = r"c:\Users\tgspc\bazi-app\bazi_backtest_results.xlsx"
wb.save(out)
print(f"Saved: {out}")
