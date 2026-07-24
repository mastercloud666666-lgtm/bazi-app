from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HDR_BG = "1A1A2E"; HDR_FG = "FFFFFF"
ALT_BG = "F9FAFB"; WIN_BG = "DCFCE7"; LOSE_BG = "FEF2F2"
POS_FG = "15803D"; NEG_FG = "DC2626"; BORDER_C = "DDDDDD"

def bd(): return Border(*[Side(style='thin', color=BORDER_C)]*0,
    left=Side(style='thin', color=BORDER_C),
    right=Side(style='thin', color=BORDER_C),
    top=Side(style='thin', color=BORDER_C),
    bottom=Side(style='thin', color=BORDER_C))
def fl(h): return PatternFill("solid", fgColor=h)
def cn(): return Alignment(horizontal='center', vertical='center')
def lf(): return Alignment(horizontal='left',   vertical='center')
def fn(bold=False, color="000000", sz=10): return Font(name='微软雅黑', bold=bold, color=color, size=sz)

GEJU_COLOR = {"中和":("1D4ED8","DBEAFE"),"身弱":("92400E","FEF9C3"),
              "身强":("B91C1C","FEE2E2"),"从儿格":("7C3AED","F3E8FF"),
              "从财格":("7C3AED","F3E8FF"),"从官格":("7C3AED","F3E8FF")}

ALL_DATA = [
    # name, gender, day_tg, geju, xi,  ar50,   ar0,   bmr,  alpha50, alpha0,  mdd50,  mdd0,  sh50,  sh0
    ("啊",         "男","戊","从儿格","金、水",  0.023, 0.024, 0.022,  0.000,  0.002, -0.256,-0.256, 0.25, 0.27),
    ("Cq女",       "女","辛","身弱",  "土、金",  0.032, 0.028, 0.022,  0.010,  0.006, -0.256,-0.256, 0.32, 0.29),
    ("程钰",       "女","丙","身弱",  "木、火",  0.023, 0.017, 0.022,  0.001, -0.006, -0.216,-0.197, 0.30, 0.26),
    ("程远",       "男","甲","身弱",  "水、木",  0.002,-0.001, 0.022, -0.020, -0.023, -0.117,-0.079, 0.08,-0.07),
    ("程宇宸",     "女","辛","中和",  "木、火、土",0.040,0.040,0.022,  0.018,  0.018, -0.289,-0.289, 0.37, 0.38),
    ("传奇妹夫",   "男","己","身弱",  "火、土",  0.019, 0.006, 0.022, -0.003, -0.016, -0.115,-0.079, 0.38, 0.19),
    ("CPR",        "女","甲","身弱",  "水、木",  0.006, 0.000, 0.022, -0.016, -0.022, -0.154,-0.125, 0.15, 0.01),
    ("Cq男",       "男","壬","身强",  "火、土",  0.028, 0.020, 0.022,  0.005, -0.002, -0.226,-0.226, 0.32, 0.26),
    ("代",         "男","壬","中和",  "火、土、金",0.041,0.042,0.022,  0.019,  0.019, -0.287,-0.285, 0.37, 0.37),
    ("戴常",       "男","癸","身弱",  "金、水",  0.011, 0.009, 0.022, -0.012, -0.013, -0.218,-0.200, 0.18, 0.17),
    ("戴元媛",     "女","己","身弱",  "火、土",  0.028, 0.021, 0.022,  0.006, -0.001, -0.170,-0.137, 0.37, 0.35),
    ("大珂国学",   "女","丙","中和",  "金、水、木",0.042,0.042,0.022,  0.020,  0.020, -0.289,-0.289, 0.37, 0.38),
    ("大侄子对象", "女","庚","中和",  "木、火、土",0.039,0.038,0.022,  0.017,  0.016, -0.286,-0.284, 0.37, 0.37),
    ("大侄子毅",   "男","庚","中和",  "木、火、土",0.041,0.041,0.022,  0.019,  0.019, -0.289,-0.289, 0.37, 0.37),
    ("东东",       "男","戊","中和",  "水、木、火",0.042,0.042,0.022,  0.020,  0.020, -0.277,-0.272, 0.38, 0.38),
    ("Fl",         "男","戊","中和",  "水、木、火",0.042,0.043,0.022,  0.020,  0.021, -0.269,-0.257, 0.38, 0.39),
    ("郭艾伦",     "男","己","从财格","水、木",  0.005,-0.002, 0.022, -0.017, -0.024, -0.195,-0.158, 0.11,-0.01),
    ("黄日华",     "女","甲","从儿格","火、土",  0.029, 0.022, 0.022,  0.007,  0.000, -0.255,-0.231, 0.32, 0.27),
    ("胡姐",       "女","癸","身弱",  "金、水",  0.020, 0.013, 0.022, -0.002, -0.010, -0.247,-0.247, 0.27, 0.20),
    ("九龙道长",   "男","丁","中和",  "金、水、木",0.042,0.042,0.022,  0.020,  0.020, -0.289,-0.289, 0.37, 0.38),
    ("Kylie",      "女","癸","从官格","土、火",  0.031, 0.024, 0.022,  0.009,  0.001, -0.226,-0.187, 0.39, 0.37),
    ("李苏涛",     "男","壬","身弱",  "金、水",  0.013, 0.008, 0.022, -0.009, -0.014, -0.247,-0.247, 0.19, 0.14),
    ("柳忠祥",     "男","癸","身强",  "火、土",  0.015, 0.010, 0.022, -0.007, -0.012, -0.158,-0.151, 0.29, 0.23),
    ("Lxt",        "女","壬","中和",  "火、土、金",0.041,0.042,0.022,  0.019,  0.019, -0.287,-0.285, 0.37, 0.37),
    ("马为",       "女","己","身弱",  "火、土",  0.028, 0.017, 0.022,  0.006, -0.005, -0.170,-0.137, 0.38, 0.30),
    ("么么撒",     "女","乙","身强",  "土、金",  0.029, 0.023, 0.022,  0.007,  0.001, -0.210,-0.177, 0.35, 0.33),
    ("宋丽娜",     "女","丙","中和",  "金、水、木",0.041,0.041,0.022,  0.019,  0.019, -0.289,-0.289, 0.37, 0.37),
    ("孙洁",       "女","丙","身弱",  "木、火",  0.028, 0.024, 0.022,  0.006,  0.002, -0.216,-0.197, 0.35, 0.35),
    ("孙平恒大",   "女","癸","中和",  "火、土、金",0.034,0.030,0.022,  0.012,  0.008, -0.254,-0.231, 0.34, 0.33),
    ("孙倩倩",     "女","辛","身弱",  "土、金",  0.023, 0.018, 0.022,  0.001, -0.004, -0.197,-0.160, 0.30, 0.28),
    ("谭鑫龙",     "男","丙","身强",  "金、水",  0.028, 0.028, 0.022,  0.006,  0.006, -0.256,-0.256, 0.31, 0.32),
    ("特朗普",     "男","己","身强",  "水、木",  0.009, 0.008, 0.022, -0.013, -0.015, -0.162,-0.133, 0.18, 0.18),
    ("Tsq",        "男","壬","中和",  "火、土、金",0.041,0.042,0.022,  0.019,  0.019, -0.287,-0.285, 0.37, 0.37),
    ("小沈阳",     "男","乙","从儿格","火、土",  0.038, 0.037, 0.022,  0.016,  0.015, -0.226,-0.226, 0.40, 0.41),
    ("圆圆爸爸",   "男","丁","中和",  "金、水、木",0.042,0.042,0.022,  0.020,  0.020, -0.289,-0.289, 0.37, 0.38),
    ("圆圆妈妈",   "女","丙","身弱",  "木、火",  0.016, 0.011, 0.022, -0.006, -0.011, -0.199,-0.160, 0.28, 0.24),
    ("圆圆姐",     "女","癸","从官格","土、火",  0.033, 0.030, 0.022,  0.011,  0.008, -0.255,-0.231, 0.36, 0.36),
    ("圆圆姨夫",   "男","壬","中和",  "火、土、金",0.034,0.030,0.022,  0.012,  0.008, -0.254,-0.231, 0.34, 0.33),
    ("张富跃",     "男","庚","身弱",  "土、金",  0.031, 0.022, 0.022,  0.009,  0.000, -0.197,-0.160, 0.38, 0.33),
    ("张丽",       "女","庚","身弱",  "土、金",  0.018, 0.013, 0.022, -0.004, -0.009, -0.179,-0.152, 0.26, 0.25),
    ("张松",       "女","庚","中和",  "木、火、土",0.041,0.041,0.022,  0.019,  0.019, -0.289,-0.289, 0.37, 0.37),
    ("张文",       "男","壬","身弱",  "金、水",  0.006, 0.005, 0.022, -0.016, -0.017, -0.189,-0.153, 0.14, 0.14),
    ("张跃芳",     "女","丁","中和",  "金、水、木",0.041,0.041,0.022,  0.019,  0.019, -0.289,-0.289, 0.37, 0.37),
    ("郑会杰",     "男","戊","中和",  "水、木、火",0.042,0.043,0.022,  0.020,  0.021, -0.269,-0.257, 0.38, 0.39),
    ("钟声",       "女","庚","中和",  "木、火、土",0.040,0.040,0.022,  0.018,  0.018, -0.289,-0.289, 0.37, 0.38),
    ("周勃舒",     "男","辛","身弱",  "土、金",  0.027, 0.022, 0.022,  0.005,  0.000, -0.139,-0.128, 0.37, 0.39),
    ("周顾",       "男","丙","从儿格","土、金",  0.037, 0.034, 0.022,  0.015,  0.012, -0.222,-0.220, 0.37, 0.36),
]

def build_sheet(wb, sheet_name, title_desc, note, ar_col, alpha_col, mdd_col, sh_col):
    ws = wb.active if sheet_name == wb.sheetnames[0] else wb.create_sheet(sheet_name)
    ws.title = sheet_name

    # 标题
    ws.merge_cells("A1:L1")
    ws["A1"] = f"八字量化回测  |  2015-01-01 ~ 2025-12-31  |  沪深300ETF（510300）  |  {title_desc}"
    ws["A1"].font = Font(name='微软雅黑', bold=True, color="1A1A2E", size=12)
    ws["A1"].fill = fl("EFF6FF")
    ws["A1"].alignment = cn()
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:L2")
    ws["A2"] = note
    ws["A2"].font = Font(name='微软雅黑', color="555555", size=9)
    ws["A2"].fill = fl("F8FAFF")
    ws["A2"].alignment = cn()
    ws.row_dimensions[2].height = 16

    ws.row_dimensions[3].height = 8  # spacer

    # 表头（可排序 → AutoFilter）
    headers = ["#","命主","性别","日主","格局","喜用","策略年化","基准年化","超额","最大回撤","夏普","跑赢"]
    ws.row_dimensions[4].height = 22
    for ci, hdr in enumerate(headers, 1):
        c = ws.cell(row=4, column=ci, value=hdr)
        c.font = Font(name='微软雅黑', bold=True, color=HDR_FG, size=10)
        c.fill = fl(HDR_BG)
        c.alignment = cn()
        c.border = bd()

    # 启用自动筛选/排序（覆盖表头行+数据区）
    ws.auto_filter.ref = f"A4:L{4 + len(ALL_DATA)}"

    for i, row in enumerate(ALL_DATA):
        r = i + 5
        name, gender, day_tg, geju, xi = row[0], row[1], row[2], row[3], row[4]
        ar     = row[ar_col]
        bmr    = row[6]   # always col index 6
        alpha  = row[alpha_col]
        mdd    = row[mdd_col]
        sharpe = row[sh_col]
        win    = alpha > 0
        even   = (i % 2 == 0)
        base   = "FFFFFF" if not even else ALT_BG
        row_bg = WIN_BG if win else (LOSE_BG if alpha < 0 else base)

        ws.row_dimensions[r].height = 18

        def wc(ci, val, nf=None, bold=False, color="000000", bg=None, al=None):
            c = ws.cell(row=r, column=ci, value=val)
            c.font = fn(bold=bold, color=color)
            c.fill = fl(bg or base)
            c.alignment = al or cn()
            c.border = bd()
            if nf: c.number_format = nf

        wc(1,  i+1, color="888888")
        wc(2,  name, bold=True, color="1A1A2E", al=lf())
        wc(3,  gender)
        wc(4,  day_tg, bold=True)

        # 格局
        gc, gb = GEJU_COLOR.get(geju, ("333333", base))
        c = ws.cell(row=r, column=5, value=geju)
        c.font = Font(name='微软雅黑', bold=True, color=gc, size=9)
        c.fill = fl(gb); c.alignment = cn(); c.border = bd()

        wc(6,  xi, al=lf())
        ar_c = POS_FG if ar >= 0.022 else (NEG_FG if ar < 0.015 else "000000")
        wc(7,  ar,     nf="0.0%", bold=True, color=ar_c)
        wc(8,  bmr,    nf="0.0%", color="555555")
        al_c = POS_FG if alpha > 0 else (NEG_FG if alpha < 0 else "888888")
        wc(9,  alpha,  nf="+0.0%;-0.0%;0.0%", bold=True, color=al_c, bg=row_bg)
        wc(10, mdd,    nf="0.0%", color=NEG_FG)
        sh_c = POS_FG if sharpe >= 0.35 else ("000000" if sharpe >= 0.2 else NEG_FG)
        wc(11, sharpe, nf="0.00", color=sh_c)

        c = ws.cell(row=r, column=12, value="✅" if win else ("—" if alpha==0 else "❌"))
        c.font = Font(name='微软雅黑', size=11)
        c.fill = fl(WIN_BG if win else (LOSE_BG if alpha < 0 else base))
        c.alignment = cn(); c.border = bd()

    # 汇总行
    last = 5 + len(ALL_DATA)
    ws.row_dimensions[last].height = 20
    ws.merge_cells(f"A{last}:F{last}")
    ws[f"A{last}"] = "平均"
    ws[f"A{last}"].font = Font(name='微软雅黑', bold=True, color=HDR_FG, size=10)
    ws[f"A{last}"].fill = fl(HDR_BG); ws[f"A{last}"].alignment = cn(); ws[f"A{last}"].border = bd()
    for ci, nf, formula in [
        (7,  "0.0%",             f"=AVERAGE(G5:G{last-1})"),
        (8,  "0.0%",             f"=AVERAGE(H5:H{last-1})"),
        (9,  "+0.0%;-0.0%;0.0%", f"=AVERAGE(I5:I{last-1})"),
        (10, "0.0%",             f"=AVERAGE(J5:J{last-1})"),
        (11, "0.00",             f"=AVERAGE(K5:K{last-1})"),
    ]:
        c = ws.cell(row=last, column=ci, value=formula)
        c.number_format = nf
        c.font = Font(name='微软雅黑', bold=True, color=HDR_FG)
        c.fill = fl(HDR_BG); c.alignment = cn(); c.border = bd()

    beat_f = f'=COUNTIF(L5:L{last-1},"✅")'
    ws.cell(row=last, column=12,
            value=f'={beat_f}&"/47 ("&TEXT({beat_f}/47,"0%")&")"'
    ).font = Font(name='微软雅黑', bold=True, color=HDR_FG)
    ws.cell(row=last, column=12).fill = fl(HDR_BG)
    ws.cell(row=last, column=12).alignment = cn()
    ws.cell(row=last, column=12).border = bd()

    # 列宽
    for ci, w in enumerate([4,13,5,5,9,14,9,9,9,9,7,9], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A5"
    return ws


# ── 文件1：纯评分仓位（无底仓） ──────────────────────────
wb1 = Workbook()
build_sheet(wb1, "纯评分仓位",
    "纯评分仓位（无底仓）",
    "仓位阶梯：评分≥70=100%  55~69=60%  45~54=30%  <45=空仓  |  大盘过滤：沪深300>120日均线",
    ar_col=6, alpha_col=9, mdd_col=11, sh_col=13)
out1 = r"c:\Users\tgspc\bazi-app\bazi_backtest_纯评分.xlsx"
wb1.save(out1)
print(f"Saved: {out1}")

# ── 文件2：底仓50% ────────────────────────────────────────
wb2 = Workbook()
build_sheet(wb2, "底仓50%",
    "底仓50%（评分加减仓）",
    "仓位阶梯：评分≥70=100%  55~69=75%  45~54=50%(底仓)  35~44=25%  <35=空仓  |  大盘过滤：沪深300>120日均线",
    ar_col=5, alpha_col=8, mdd_col=10, sh_col=12)
out2 = r"c:\Users\tgspc\bazi-app\bazi_backtest_底仓50.xlsx"
wb2.save(out2)
print(f"Saved: {out2}")
